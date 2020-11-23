import os
import openpyxl


#input is an excel file, each hz in a different sheet. 
#first column=the offset, subsequent columns are residue intensity data.
#make sure fq3list from run folder is in the same directory

file = input('File Name (.xlsx format)>>>')
#file='3mM_Compiled_11-18.xlsx'
error = input('Estimated Error>>>')
#error= '1E4'

#setting up the offset list
fqlist = [-100000]
f = open('fq3list','r')
for line in f:
   if 'ppm' not in line:
      ppm_shift = float(line.strip('\n'))
      frq_shift = ppm_shift * 81.094         #make sure these numbers make sense!
      frq_shift_adjusted = frq_shift - 9486.9  #they're used for N on the 800MHz
      fqlist.append(str(frq_shift_adjusted))
f.close()


wb=openpyxl.load_workbook(file)

sheetnames = wb.get_sheet_names()
residue_master_list=[]
for a in sheetnames:
   #open each sheet
   sheet = wb.get_sheet_by_name(a)

   #set up list of residues to use later
   residue_list=[]

   #iterate over residues and save files as residue#.txt
   for column in sheet.iter_cols(min_col=2, values_only=True):
      residue=str(column[0])
      filename= str(residue) + '.txt'
      filepath= os.path.join(a,filename)
      if not os.path.isdir(a):
         os.mkdir(a)
      f = open(filepath, 'w')
      for i in range(len(column)-1):
         f.write(str(fqlist[i]) + '\t'+ str(column[i+1]) + '\t' + error + '\n')
      f.close()
      #collect residues in a list for later
      residue_list.append(residue)
   #hack because I don't know how to make it do it only once :P   
   residue_master_list.append(residue_list)

residues=residue_master_list[0]
#prep the input files. you'll need to fill in the chemical shift and dw_ab values by hand, but at least the residues are there
cs_a=open('cs_a', 'w')
dw_ab=open('dw_ab','w')
experiment=open('experiment','w')
for n in residues:   
   cs_a.write(n+'N'+'\n')
   dw_ab.write(n+'N'+'\n')
   #copy this bit into the bottom of the 'experiment' files rather than typing by hand
   experiment.write(n+'N-HN = '+n+'.txt'+'\n')
cs_a.close()
dw_ab.close()  
experiment.close()   

    